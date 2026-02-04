"""
Generatore di lista tagli (cut list) per ottimizzazione produzione
Calcola dimensioni nette e listarelle necessarie per ogni pannello
"""

import adsk.core
import adsk.fusion
from collections import defaultdict

class CutList:
    """Generatore di lista tagli da componenti Fusion"""
    
    def __init__(self, component):
        """
        Inizializza il generatore di cutlist
        
        Args:
            component: Componente Fusion da analizzare
        """
        self.component = component
        self.parts = []
    
    def generate(self, include_hardware=False):
        """
        Genera la lista tagli dal componente
        
        Args:
            include_hardware: Include ferramenta nella lista (default False)
        
        Returns:
            dict: Lista tagli organizzata per materiale e dimensioni
        """
        self.parts = []
        
        # Analizza tutti i corpi del componente
        self._analyze_bodies(self.component)
        
        # Analizza sotto-componenti
        for occurrence in self.component.occurrences:
            self._analyze_bodies(occurrence.component)
        
        # Organizza per materiale e dimensioni
        organized = self._organize_parts()
        
        # Calcola statistiche
        stats = self._calculate_statistics(organized)
        
        return {
            'parts': organized,
            'statistics': stats,
            'total_parts': len(self.parts)
        }
    
    def _analyze_bodies(self, component):
        """
        Analizza i corpi di un componente
        
        Args:
            component: Componente da analizzare
        """
        for body in component.bRepBodies:
            if not body.isVisible:
                continue
            
            part_info = self._extract_part_info(body)
            if part_info:
                self.parts.append(part_info)
    
    def _extract_part_info(self, body):
        """
        Estrae informazioni da un corpo
        
        Args:
            body: BRepBody da analizzare
        
        Returns:
            dict: Informazioni parte o None
        """
        try:
            # Ottieni bounding box
            bbox = body.boundingBox
            
            # Calcola dimensioni (converti cm a mm)
            width = (bbox.maxPoint.x - bbox.minPoint.x) * 10
            height = (bbox.maxPoint.y - bbox.minPoint.y) * 10
            depth = (bbox.maxPoint.z - bbox.minPoint.z) * 10
            
            # Ordina le dimensioni per identificare lo spessore
            dims = sorted([width, height, depth])
            
            # Lo spessore è la dimensione più piccola
            thickness = dims[0]
            length = dims[2]
            width_net = dims[1]
            
            # Determina se è un pannello (dimensione più piccola < 30mm tipicamente)
            is_panel = thickness < 30
            
            if not is_panel:
                return None
            
            # Ottieni materiale (se assegnato)
            material_name = "Non Assegnato"
            if body.material:
                material_name = body.material.name
            
            # Calcola area
            area = length * width_net / 1000000  # mm² to m²
            
            # Determina tipo bordi (semplificato)
            edge_bands = self._determine_edge_bands(body.name)
            
            return {
                'name': body.name,
                'length': round(length, 1),
                'width': round(width_net, 1),
                'thickness': round(thickness, 1),
                'material': material_name,
                'area': round(area, 4),
                'edge_bands': edge_bands,
                'quantity': 1
            }
        except:
            return None
    
    def _determine_edge_bands(self, body_name):
        """
        Determina quali bordi necessitano listarella
        
        Args:
            body_name: Nome del corpo
        
        Returns:
            dict: Bordi da listarellare
        """
        # Logica semplificata basata sul nome
        edges = {
            'front': False,
            'back': False,
            'left': False,
            'right': False
        }
        
        # I pannelli visibili tipicamente hanno bordo frontale
        if 'Anta' in body_name or 'Frontale' in body_name:
            edges['front'] = True
            edges['back'] = False
            edges['left'] = True
            edges['right'] = True
        elif 'Fianco' in body_name:
            edges['front'] = True
        elif 'Ripiano' in body_name:
            edges['front'] = True
        
        return edges
    
    def _organize_parts(self):
        """
        Organizza le parti per materiale e dimensioni
        
        Returns:
            dict: Parti organizzate
        """
        organized = defaultdict(lambda: defaultdict(list))
        
        for part in self.parts:
            material = part['material']
            thickness = part['thickness']
            
            # Cerca se esiste già una parte con stesse dimensioni
            found = False
            for existing in organized[material][thickness]:
                if (existing['length'] == part['length'] and 
                    existing['width'] == part['width']):
                    existing['quantity'] += 1
                    existing['names'].append(part['name'])
                    found = True
                    break
            
            if not found:
                part['names'] = [part['name']]
                organized[material][thickness].append(part)
        
        # Converti defaultdict in dict normale
        result = {}
        for material, thicknesses in organized.items():
            result[material] = {}
            for thickness, parts in thicknesses.items():
                result[material][thickness] = parts
        
        return result
    
    def _calculate_statistics(self, organized):
        """
        Calcola statistiche sulla lista tagli
        
        Args:
            organized: Parti organizzate
        
        Returns:
            dict: Statistiche
        """
        stats = {
            'total_area': 0,
            'by_material': {},
            'by_thickness': defaultdict(float)
        }
        
        for material, thicknesses in organized.items():
            material_area = 0
            
            for thickness, parts in thicknesses.items():
                for part in parts:
                    part_area = part['area'] * part['quantity']
                    material_area += part_area
                    stats['by_thickness'][thickness] += part_area
            
            stats['by_material'][material] = round(material_area, 4)
            stats['total_area'] += material_area
        
        stats['total_area'] = round(stats['total_area'], 4)
        stats['by_thickness'] = dict(stats['by_thickness'])
        
        return stats
    
    def export_to_csv(self, filepath):
        """
        Esporta la lista tagli in formato CSV
        
        Args:
            filepath: Path del file di output
        
        Returns:
            bool: Successo operazione
        """
        try:
            import csv
            
            cutlist = self.generate()
            
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                
                # Header
                writer.writerow([
                    'Materiale', 'Spessore', 'Lunghezza', 'Larghezza',
                    'Quantità', 'Area m²', 'Bordo Fronte', 'Bordo Retro',
                    'Bordo Sx', 'Bordo Dx', 'Nome'
                ])
                
                # Dati
                for material, thicknesses in cutlist['parts'].items():
                    for thickness, parts in thicknesses.items():
                        for part in parts:
                            edges = part['edge_bands']
                            writer.writerow([
                                material,
                                thickness,
                                part['length'],
                                part['width'],
                                part['quantity'],
                                part['area'],
                                'Sì' if edges.get('front') else 'No',
                                'Sì' if edges.get('back') else 'No',
                                'Sì' if edges.get('left') else 'No',
                                'Sì' if edges.get('right') else 'No',
                                ', '.join(part['names'])
                            ])
                
                # Statistiche
                writer.writerow([])
                writer.writerow(['STATISTICHE'])
                writer.writerow(['Area Totale (m²)', cutlist['statistics']['total_area']])
                writer.writerow([])
                writer.writerow(['Per Materiale'])
                for material, area in cutlist['statistics']['by_material'].items():
                    writer.writerow([material, area])
            
            return True
        except Exception as e:
            print(f"❌ Errore export CSV: {e}")
            return False
    
    def export_to_excel(self, filepath):
        """
        Esporta la lista tagli in formato Excel
        
        Args:
            filepath: Path del file di output
        
        Returns:
            bool: Successo operazione
        """
        try:
            # Richiede openpyxl (installazione opzionale)
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            cutlist = self.generate()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Lista Tagli"
            
            # Header con formattazione
            headers = [
                'Materiale', 'Spessore', 'Lunghezza', 'Larghezza',
                'Quantità', 'Area m²', 'Listarelle', 'Nome'
            ]
            
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(1, col, header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Dati
            row = 2
            for material, thicknesses in cutlist['parts'].items():
                for thickness, parts in thicknesses.items():
                    for part in parts:
                        edges = part['edge_bands']
                        edge_str = []
                        if edges.get('front'): edge_str.append('F')
                        if edges.get('back'): edge_str.append('R')
                        if edges.get('left'): edge_str.append('S')
                        if edges.get('right'): edge_str.append('D')
                        
                        ws.cell(row, 1, material)
                        ws.cell(row, 2, thickness)
                        ws.cell(row, 3, part['length'])
                        ws.cell(row, 4, part['width'])
                        ws.cell(row, 5, part['quantity'])
                        ws.cell(row, 6, part['area'])
                        ws.cell(row, 7, ', '.join(edge_str) if edge_str else 'Nessuno')
                        ws.cell(row, 8, ', '.join(part['names']))
                        
                        row += 1
            
            # Salva
            wb.save(filepath)
            return True
        except ImportError:
            print("⚠️ openpyxl non installato - usa export CSV")
            return False
        except Exception as e:
            print(f"❌ Errore export Excel: {e}")
            return False
